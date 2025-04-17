import openpyxl
import os

DATA_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "data")
EXCEL_FILE = os.path.join(DATA_DIR, "datos_api.xlsx")

os.makedirs(DATA_DIR, exist_ok=True)

def create_excel(headers):
    if os.path.exists(EXCEL_FILE):
        return False
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Datos"
    ws.append(headers)
    wb.save(EXCEL_FILE)
    return True

def insert_row(values):
    if not os.path.exists(EXCEL_FILE):
        return False
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append(values)
    wb.save(EXCEL_FILE)
    return True

def update_row(row_idx, values):
    if not os.path.exists(EXCEL_FILE):
        return False
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    if row_idx < 2 or row_idx > ws.max_row:
        return False
    for col, value in enumerate(values, start=1):
        ws.cell(row=row_idx, column=col, value=value)
    wb.save(EXCEL_FILE)
    return True

def delete_row(row_idx):
    if not os.path.exists(EXCEL_FILE):
        return False
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    if row_idx < 2 or row_idx > ws.max_row:
        return False
    ws.delete_rows(row_idx)
    wb.save(EXCEL_FILE)
    return True

def get_data():
    if not os.path.exists(EXCEL_FILE):
        return []
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    return [[cell.value for cell in row] for row in ws.iter_rows(values_only=True)]