import tkinter as tk
import random
import openpyxl  # Biblioteca para manejar archivos Excel
import os  # Importar el módulo os
import matplotlib.pyplot as plt
from collections import Counter
import math

# Constantes
EXCEL_FILE = "numeros_generados.xlsx"
#TOMBOLA_TITULOS = ["TOMBOLLA PRINCIPAL (1-40)", "BOLA BONO 1 (1-12)", "BOLA BONO 2 (1-15)"]
COLORS = {
    "background": "#4CAF50",
    "button": "#FF9800",
    "highlight": "#f44336",
    "text": "white",
    "default": "black"
}

class Tombola:
    """
    Clase que representa una tómbola con números generados aleatoriamente.
    """

    def __init__(self, master, rango, columnas, max_sorteos, titulo):
        """
        Inicializa la tómbola con los parámetros dados.

        Args:
            master (tk.Frame): El marco principal donde se colocará la tómbola.
            rango (tuple): Rango de números (min, max).
            columnas (int): Número de columnas para mostrar los números.
            max_sorteos (int): Número máximo de sorteos permitidos.
            titulo (str): Título de la tómbola.
        """
        self.master = master
        self.min_num, self.max_num = rango
        self.columnas = columnas
        self.max_sorteos = max_sorteos
        self.titulo = titulo

        self.numeros = list(range(self.min_num, self.max_num + 1))
        self.sacados = []
        self.prediccion = None
        self.prediccion_aciertos = 0

        self.crear_widgets()
        self.actualizar_bombo()

    def guardar_en_excel(self):
        """
        Guarda los números generados en un archivo Excel, todos en la misma hoja.
        """
        ruta_actual = os.path.dirname(os.path.abspath(__file__))
        archivo = os.path.join(ruta_actual, EXCEL_FILE)

        try:
            wb = openpyxl.load_workbook(archivo)
        except FileNotFoundError:
            wb = openpyxl.Workbook()

        if "Resultados" not in wb.sheetnames:
            ws = wb.create_sheet("Resultados")
            ws.append(["Posición 1", "Posición 2", "Posición 3", "Posición 4", "Posición 5", "Posición 6", "BOLA BONO 1", "BOLA BONO 2"])
        else:
            ws = wb["Resultados"]

        fila = ws.max_row + 1
        if self.titulo == "TOMBOLLA PRINCIPAL (1-40)":
            for i, numero in enumerate(self.sacados):
                ws.cell(row=fila, column=i + 1, value=numero)
        elif self.titulo == "BOLA BONO 1 (1-12)" and self.sacados:
            ws.cell(row=fila - 1, column=7, value=self.sacados[0])
        elif self.titulo == "BOLA BONO 2 (1-15)" and self.sacados:
            ws.cell(row=fila - 1, column=8, value=self.sacados[0])

        wb.save(archivo)

    def predecir_en_base_a_excel(self):
        """
        Predice un número basado en los datos del archivo Excel.
        """
        ruta_actual = os.path.dirname(os.path.abspath(__file__))
        archivo = os.path.join(ruta_actual, EXCEL_FILE)

        try:
            wb = openpyxl.load_workbook(archivo)
            if "Resultados" in wb.sheetnames:
                ws = wb["Resultados"]
                numeros_generados = [
                    num for row in ws.iter_rows(min_row=2, max_col=8, values_only=True) for num in row if num is not None
                ]

                if numeros_generados:
                    conteo = Counter(numeros_generados)
                    menos_frecuentes = sorted(conteo.items(), key=lambda x: x[1])[:5]
                    prediccion = random.choice([num for num, _ in menos_frecuentes])
                    self.ultimo_numero.set(f"Predicción basada en Excel: {prediccion}")
                    return
        except FileNotFoundError:
            pass

        self.ultimo_numero.set("No hay datos en Excel para predecir")

    def predecir_posicion(self, posicion):
        """Predice un número para una posición específica."""
        ruta_actual = os.path.dirname(os.path.abspath(__file__))
        archivo = os.path.join(ruta_actual, "numeros_generados.xlsx")
        try:
            wb = openpyxl.load_workbook(archivo)
            if "Resultados" in wb.sheetnames:
                ws = wb["Resultados"]
                numeros_posicion = [
                    row[posicion - 1] for row in ws.iter_rows(min_row=2, max_col=8, values_only=True)
                    if row[posicion - 1] is not None
                ]
                if numeros_posicion:
                    # Predicción basada en los números menos frecuentes
                    from collections import Counter
                    conteo = Counter(numeros_posicion)
                    menos_frecuentes = sorted(conteo.items(), key=lambda x: x[1])[:5]
                    prediccion = random.choice([num for num, _ in menos_frecuentes])
                    self.ultimo_numero.set(f"Predicción posición {posicion}: {prediccion}")
                    return
        except FileNotFoundError:
            pass
        
        self.ultimo_numero.set(f"No hay datos para predecir posición {posicion}")

    def analizar_numeros(self):
        """Analiza los números calientes, fríos y con mayor probabilidad."""
        ruta_actual = os.path.dirname(os.path.abspath(__file__))
        archivo = os.path.join(ruta_actual, "numeros_generados.xlsx")
        try:
            wb = openpyxl.load_workbook(archivo)
            if "Resultados" in wb.sheetnames:
                ws = wb["Resultados"]
                numeros_generados = []
                for row in ws.iter_rows(min_row=2, max_col=8, values_only=True):
                    numeros_generados.extend(filter(None, row))  # Agregar todos los números no nulos
                
                if numeros_generados:
                    # Contar la frecuencia de cada número
                    from collections import Counter
                    conteo = Counter(numeros_generados)
                    calientes = sorted(conteo.items(), key=lambda x: x[1], reverse=True)[:5]  # Top 5 más frecuentes
                    frios = sorted(conteo.items(), key=lambda x: x[1])[:5]  # Top 5 menos frecuentes
                    
                    # Mostrar el análisis en pantalla
                    resultado = f"Calientes: {', '.join([str(num) for num, _ in calientes])}\n"
                    resultado += f"Fríos: {', '.join([str(num) for num, _ in frios])}"
                    self.texto_analisis.config(state=tk.NORMAL)
                    self.texto_analisis.delete(1.0, tk.END)
                    self.texto_analisis.insert(tk.END, resultado)
                    self.texto_analisis.config(state=tk.DISABLED)
                    return
        except FileNotFoundError:
            pass
    
        self.texto_analisis.config(state=tk.NORMAL)
        self.texto_analisis.delete(1.0, tk.END)
        self.texto_analisis.insert(tk.END, "No hay datos para análisis")
        self.texto_analisis.config(state=tk.DISABLED)
    def _mostrar_analisis(self,resultado):
        """Muestra el análisis de números en el Text widget."""

        self.texto_analisis.config(state=tk.NORMAL)
        self.texto_analisis.delete(1.0, tk.END)
        self.texto_analisis.insert(tk.END, resultado)
        if resultado:
            self.texto_analisis.config(state=tk.DISABLED)
        else:       
            self.texto_analisis.config(state=tk.NORMAL)
            self.texto_analisis.delete(1.0, tk.END)
       
        self.texto_analisis.config(state=tk.DISABLED)    

    def analizar_numeros_principal(self):
        """Analiza los números calientes y fríos para TOMBOLLA PRINCIPAL (1-40)."""
        ruta_actual = os.path.dirname(os.path.abspath(__file__))
        archivo = os.path.join(ruta_actual, "numeros_generados.xlsx")
        try:
            wb = openpyxl.load_workbook(archivo)
            if "Resultados" in wb.sheetnames:
                ws = wb["Resultados"]
                numeros_generados = []
                for row in ws.iter_rows(min_row=2, max_col=6, values_only=True):  # Columnas 1 a 6
                    numeros_generados.extend(filter(None, row))  # Agregar todos los números no nulos
                
                if numeros_generados:
                    # Contar la frecuencia de cada número
                    from collections import Counter
                    conteo = Counter(numeros_generados)
                    calientes = sorted(conteo.items(), key=lambda x: x[1], reverse=True)[:5]  # Top 5 más frecuentes
                    frios = sorted(conteo.items(), key=lambda x: x[1])[:5]  # Top 5 menos frecuentes
                    
                    # Mostrar el análisis en pantalla
                    resultado = f"Calientes (Principal): {', '.join([str(num) for num, _ in calientes])}\n"
                    resultado += f"Fríos (Principal): {', '.join([str(num) for num, _ in frios])}"
                    self._mostrar_analisis(resultado)
                    return
        except FileNotFoundError:
            pass
        
        self._mostrar_analisis("No hay datos para análisis en TOMBOLLA PRINCIPAL")

    def analizar_numeros_bono1(self):
        """Analiza los números calientes y fríos para BOLA BONO 1."""
        ruta_actual = os.path.dirname(os.path.abspath(__file__))
        archivo = os.path.join(ruta_actual, "numeros_generados.xlsx")
        try:
            wb = openpyxl.load_workbook(archivo)
            if "Resultados" in wb.sheetnames:
                ws = wb["Resultados"]
                numeros_generados = [
                    row[6] for row in ws.iter_rows(min_row=2, max_col=7, values_only=True) if row[6] is not None
                ]  # Columna 7 para BOLA BONO 1
                
                if numeros_generados:
                    # Contar la frecuencia de cada número
                    from collections import Counter
                    conteo = Counter(numeros_generados)
                    calientes = sorted(conteo.items(), key=lambda x: x[1], reverse=True)[:5]
                    frios = sorted(conteo.items(), key=lambda x: x[1])[:5]
                    
                    # Mostrar el análisis en pantalla
                    resultado = f"Calientes (Bono 1): {', '.join([str(num) for num, _ in calientes])}\n"
                    resultado += f"Fríos (Bono 1): {', '.join([str(num) for num, _ in frios])}"
                    self._mostrar_analisis(resultado)
                    return
        except FileNotFoundError:
            pass
        
        self._mostrar_analisis("No hay datos para análisis en BOLA BONO 1")

    def analizar_numeros_bono2(self):
        """Analiza los números calientes y fríos para BOLA BONO 2."""
        ruta_actual = os.path.dirname(os.path.abspath(__file__))
        archivo = os.path.join(ruta_actual, "numeros_generados.xlsx")
        try:
            wb = openpyxl.load_workbook(archivo)
            if "Resultados" in wb.sheetnames:
                ws = wb["Resultados"]
                numeros_generados = [
                    row[7] for row in ws.iter_rows(min_row=2, max_col=8, values_only=True) if row[7] is not None
                ]  # Columna 8 para BOLA BONO 2
                
                if numeros_generados:
                    # Contar la frecuencia de cada número
                    from collections import Counter
                    conteo = Counter(numeros_generados)
                    calientes = sorted(conteo.items(), key=lambda x: x[1], reverse=True)[:5]
                    frios = sorted(conteo.items(), key=lambda x: x[1])[:5]
                    
                    # Mostrar el análisis en pantalla
                    resultado = f"Calientes (Bono 2): {', '.join([str(num) for num, _ in calientes])}\n"
                    resultado += f"Fríos (Bono 2): {', '.join([str(num) for num, _ in frios])}"
                    self._mostrar_analisis(resultado)
                    return
        except FileNotFoundError:
            pass
        
        self._mostrar_analisis("No hay datos para análisis en BOLA BONO 2")

    def generar_todos(self):
        """
        Genera todos los números automáticamente.
        """
        while self.numeros and len(self.sacados) < self.max_sorteos:
            numero = random.choice(self.numeros)
            self.sacados.append(numero)
            self.numeros.remove(numero)

        self.sacados.sort()
        self.actualizar_bombo()
        self._actualizar_texto_sacados()
        self.ultimo_numero.set("Sorteo completo")
        self.btn_sacar.config(text="SORTEO COMPLETO", state=tk.DISABLED)
        self.btn_generar_todos.config(state=tk.DISABLED)
        self.guardar_en_excel()

    def crear_widgets(self):
        """
        Crea los widgets de la interfaz para la tómbola.
        """
        self.marco = tk.Frame(self.master, bd=2, relief=tk.GROOVE)
        self.marco.pack(pady=10, padx=10, fill=tk.BOTH)

        tk.Label(self.marco, text=self.titulo, font=("Arial", 12, "bold")).pack(pady=5)

        self.frame_bombo = tk.Frame(self.marco)
        self.frame_bombo.pack(pady=5)

        self.frame_numeros = tk.Frame(self.frame_bombo)
        self.frame_numeros.pack()

        self.btn_sacar = tk.Button(self.marco, text="Sacar número", command=self.iniciar_sorteo,
                                   font=("Arial", 10), bg=COLORS["background"], fg=COLORS["text"])
        self.btn_sacar.pack(pady=5)

        self.btn_generar_todos = tk.Button(self.marco, text="Generar todos", command=self.generar_todos,
                                           font=("Arial", 10), bg=COLORS["button"], fg=COLORS["text"])
        self.btn_generar_todos.pack(pady=5)

        self.ultimo_numero = tk.StringVar()
        tk.Label(self.marco, textvariable=self.ultimo_numero, font=("Arial", 14), fg=COLORS["highlight"]).pack(pady=5)

        self.lista_sacados = tk.Text(self.marco, width=30, height=4, wrap=tk.WORD, state=tk.DISABLED)
        self.lista_sacados.pack(pady=5)

        # Botón para predecir un número basado en Excel
        self.btn_predecir_excel = tk.Button(self.marco, text="Predecir con Excel", 
                                            command=self.predecir_en_base_a_excel,
                                            font=("Arial", 10), bg="#FFC107", fg="black")
        self.btn_predecir_excel.pack(pady=5)

        if self.titulo == "TOMBOLLA PRINCIPAL (1-40)":
        # Botones para predecir cada posición
            for i in range(1, 7):  # 6 posiciones
                btn_predecir_posicion = tk.Button(self.marco, text=f"Predecir Posición {i}", 
                                                command=lambda i=i: self.predecir_posicion(i),
                                                font=("Arial", 10), bg="#FFC107", fg="black")
                btn_predecir_posicion.pack(pady=2)

        # Botón para analizar números calientes y fríos
            btn_analizar = tk.Button(self.marco, text="Analizar Números", 
                                    command=self.analizar_numeros,
                                    font=("Arial", 10), bg="#4CAF50", fg="white")
            btn_analizar.pack(pady=5)

        # Botón para analizar TOMBOLLA PRINCIPAL
            btn_analizar_principal = tk.Button(self.marco, text="Analizar Principal", 
                                            command=self.analizar_numeros_principal,
                                            font=("Arial", 10), bg="#4CAF50", fg="white")
            btn_analizar_principal.pack(pady=5)
        
        if  self.titulo == "BOLA BONO 1 (1-12)":
        # Botón para analizar BOLA BONO 1
            btn_analizar_bono1 = tk.Button(self.marco, text="Analizar Bono 1", 
                                            command=self.analizar_numeros_bono1,
                                            font=("Arial", 10), bg="#4CAF50", fg="white")
            btn_analizar_bono1.pack(pady=5)
        if self.titulo == "BOLA BONO 2 (1-15)":
        # Botón para analizar BOLA BONO 2
            btn_analizar_bono2 = tk.Button(self.marco, text="Analizar Bono 2", 
                                            command=self.analizar_numeros_bono2,
                                            font=("Arial", 10), bg="#4CAF50", fg="white")
            btn_analizar_bono2.pack(pady=5)

        # Text widget para mostrar el análisis de números
        self.texto_analisis = tk.Text(self.marco, width=40, height=5, wrap=tk.WORD, state=tk.DISABLED)
        self.texto_analisis.pack(pady=5)

        btn_estadisticas = tk.Button(self.marco, text="Generar Estadísticas", 
                                     command=self.generar_estadisticas,
                                     font=("Arial", 10), bg="#4CAF50", fg="white")
        btn_estadisticas.pack(pady=5)

        btn_estadisticas_por_posicion = tk.Button(self.marco, text="Generar Estadísticas por Posición", 
                                                  command=self.generar_estadisticas_por_posicion,
                                                  font=("Arial", 10), bg="#4CAF50", fg="white")
        btn_estadisticas_por_posicion.pack(pady=5)

        btn_estadisticas_posicion = tk.Button(self.marco, text="Estadísticas por Posición", 
                                              command=self.estadisticas_por_posicion,
                                              font=("Arial", 10), bg="#4CAF50", fg="white")
        btn_estadisticas_posicion.pack(pady=5)

    def actualizar_bombo(self):
        """
        Actualiza los números en el bombo.
        """
        for widget in self.frame_numeros.winfo_children():
            widget.destroy()

        self.numeros.sort()
        for i, num in enumerate(self.numeros):
            fila, columna = divmod(i, self.columnas)
            lbl = tk.Label(self.frame_numeros, text=str(num), width=3, relief=tk.RIDGE,
                           font=("Arial", 10), bg="white")
            lbl.grid(row=fila, column=columna, padx=2, pady=2)

    def iniciar_sorteo(self):
        """
        Inicia el sorteo de un número.
        """
        if self.numeros and len(self.sacados) < self.max_sorteos:
            self.btn_sacar.config(state=tk.DISABLED)
            numero = random.choice(self.numeros)
            self.master.after(1000, lambda: self.completar_sorteo(numero))

    def completar_sorteo(self, numero):
        """
        Completa el sorteo y verifica si la predicción se cumplió.
        """
        if numero in self.numeros:
            self.numeros.remove(numero)
        self.sacados.append(numero)
        self.actualizar_bombo()

        if numero == self.prediccion:
            self.prediccion_aciertos += 1
            self.ultimo_numero.set(f"¡Acierto! Predicción: {self.prediccion} (Aciertos: {self.prediccion_aciertos})")
        else:
            self.ultimo_numero.set(f"Último: {numero} (Aciertos: {self.prediccion_aciertos})")

        self._actualizar_texto_sacados()

        if len(self.sacados) >= self.max_sorteos or not self.numeros:
            self.btn_sacar.config(text="SORTEO COMPLETO", state=tk.DISABLED)

    def _actualizar_texto_sacados(self):
        """
        Actualiza el widget de texto con los números sacados.
        """
        self.lista_sacados.config(state=tk.NORMAL)
        self.lista_sacados.delete(1.0, tk.END)
        columnas = 5
        for i, num in enumerate(self.sacados):
            if i > 0 and i % columnas == 0:
                self.lista_sacados.insert(tk.END, "\n")
            self.lista_sacados.insert(tk.END, f"{num:3} ")
        self.lista_sacados.config(state=tk.DISABLED)

    def reiniciar(self):
        """Reinicia la tómbola a su estado inicial."""
        self.numeros = list(range(self.min_num, self.max_num + 1))
        self.sacados = []
        self.prediccion = None
        self.prediccion_aciertos = 0
        self.actualizar_bombo()
        self.lista_sacados.config(state=tk.NORMAL)
        self.lista_sacados.delete(1.0, tk.END)  # Limpiar el widget Text
        self.lista_sacados.config(state=tk.DISABLED)
        self.ultimo_numero.set("")
        self.btn_sacar.config(text="Sacar número", state=tk.NORMAL)
        self.btn_generar_todos.config(state=tk.NORMAL)

    def predecir_numero(self):
        """Realiza una predicción de un número aleatorio de los números restantes."""
        if self.numeros:
            self.prediccion = random.choice(self.numeros)
            self.ultimo_numero.set(f"Predicción: {self.prediccion}")
        else:
            self.ultimo_numero.set("No hay números para predecir")

    def generar_estadisticas(self):
        """
        Genera estadísticas por posición para cada tómbola y muestra gráficos en una sola ventana.
        """
        ruta_actual = os.path.dirname(os.path.abspath(__file__))
        archivo = os.path.join(ruta_actual, "numeros_generados.xlsx")

        # Definir las posiciones correspondientes a cada tómbola
        posiciones_tombola = {
            "TOMBOLLA PRINCIPAL (1-40)": range(1, 7),  # Columnas 1 a 6
            "BOLA BONO 1 (1-12)": range(7, 8),        # Columna 7
            "BOLA BONO 2 (1-15)": range(8, 9)         # Columna 8
        }

        try:
            wb = openpyxl.load_workbook(archivo)
            if "Resultados" in wb.sheetnames:
                ws = wb["Resultados"]

                # Obtener las posiciones correspondientes a la tómbola actual
                posiciones = posiciones_tombola.get(self.titulo, [])
                if not posiciones:
                    self._mostrar_analisis("No hay posiciones definidas para esta tómbola.")
                    return
                
                # Crear una ventana para los gráficos
                fig, axes = plt.subplots(len(posiciones), 1, figsize=(15, 5 * len(posiciones)))
                
                if len(posiciones) == 1:
                    axes = [axes]  # Asegurar que `axes` sea iterable si hay solo un gráfico

                # Generar gráficos para cada posición
                for ax, posicion in zip(axes, posiciones):
                    numeros_posicion = [
                        row[posicion - 1] for row in ws.iter_rows(min_row=2, max_col=8, values_only=True)
                        if row[posicion - 1] is not None
                    ]

                    if numeros_posicion:
                        # Contar la frecuencia de cada número
                        from collections import Counter
                        conteo = Counter(numeros_posicion)

                        # Preparar datos para el gráfico
                        numeros = list(conteo.keys())
                        frecuencias = list(conteo.values())

                        # Crear el gráfico de barras
                        barras = ax.bar(numeros, frecuencias, color='skyblue')
                        ax.bar_label(barras, fontsize=10, padding=3,color='red')

                        ax.set_title(f"Frecuencia de Números - Posición {posicion}", fontsize=14)
                        ax.set_xlabel("Números", fontsize=12)
                        ax.set_ylabel("Frecuencia", fontsize=12)
                        ax.set_xticks(numeros)
                        ax.set_xticklabels(numeros, rotation=45)
                        ax.grid(axis='y', linestyle='--', alpha=0.7)

                plt.tight_layout(pad=12.0, h_pad=11.0)
                plt.show()
            else:
                self._mostrar_analisis("No hay datos en Excel para generar estadísticas.")
        except FileNotFoundError:
            self._mostrar_analisis("Archivo Excel no encontrado.")

    def generar_estadisticas_por_posicion(self):
        """Genera estadísticas por posición de cada número y muestra gráficos."""
        ruta_actual = os.path.dirname(os.path.abspath(__file__))
        archivo = os.path.join(ruta_actual, "numeros_generados.xlsx")
        posicion_tombola = {
        "GENERAL": (1, 7),
        "BOLA BONO 1": (7, 8),
        "BOLA BONO 2": (8, 9)
        }
        try:
            wb = openpyxl.load_workbook(archivo)
            if "Resultados" in wb.sheetnames:
                ws = wb["Resultados"]
                if self.titulo == "TOMBOLLA PRINCIPAL (1-40)":
                    position_rango = posicion_tombola["GENERAL"]
                elif self.titulo == "BOLA BONO 1 (1-12)":
                    position_rango = posicion_tombola["BOLA BONO 1"]
                elif self.titulo == "BOLA BONO 2 (1-15)":
                    position_rango = posicion_tombola["BOLA BONO 2"]
                else:   
                    position_rango = (1, 9)
                # Iterar por cada posición (columna)
                inicio, fin = position_rango
                for posicion in range( inicio, fin):  # Columnas 1 a 8
                    numeros_posicion = [
                        row[posicion - 1] for row in ws.iter_rows(min_row=2, max_col=8, values_only=True)
                        if row[posicion - 1] is not None
                    ]
                    
                    if numeros_posicion:
                        # Contar la frecuencia de cada número
                        from collections import Counter
                        conteo = Counter(numeros_posicion)
                        
                        # Preparar datos para el gráfico
                        numeros = list(conteo.keys())
                        frecuencias = list(conteo.values())
                        
                        # Crear el gráfico de barras
                        plt.figure(figsize=(10, 6))
                        plt.bar(numeros, frecuencias, color='skyblue')
                        plt.title(f"Frecuencia de Números - Posición {posicion}", fontsize=16)
                        plt.xlabel("Números", fontsize=12)
                        plt.ylabel("Frecuencia", fontsize=12)
                        plt.xticks(numeros,rotation=45)
                        plt.tight_layout()
                        
                        # Mostrar el gráfico
                        plt.show()
            else:
                self._mostrar_analisis("No hay datos en Excel para generar estadísticas por posición")
        except FileNotFoundError:
            self._mostrar_analisis("Archivo Excel no encontrado")
     
    def estadisticas_por_posicion(self):
        """Genera estadísticas por posición de cada número y muestra gráficos."""
        ruta_actual = os.path.dirname(os.path.abspath(__file__))
        archivo = os.path.join(ruta_actual, "numeros_generados.xlsx")
        posicion_tombola = {
        "GENERAL": (1, 6),
        "BOLA BONO 1": (7, 7),
        "BOLA BONO 2": (8, 8)
        }
        try:
            wb = openpyxl.load_workbook(archivo)
            if "Resultados" in wb.sheetnames:
                ws = wb["Resultados"]
                if self.titulo == "TOMBOLLA PRINCIPAL (1-40)":
                    position_rango = posicion_tombola["GENERAL"]
                elif self.titulo == "BOLA BONO 1 (1-12)":
                    position_rango = posicion_tombola["BOLA BONO 1"]
                elif self.titulo == "BOLA BONO 2 (1-15)":
                    position_rango = posicion_tombola["BOLA BONO 2"]
                else:   
                    position_rango = (1, 9)
                # Iterar por cada posición (columna)
                inicio, fin = position_rango
               # Calcular cuántos gráficos vamos a hacer
                cantidad_graficos = fin - inicio + 1

                # Calcular filas y columnas para distribuir bien los subplots
                filas = (cantidad_graficos // 3) + (1 if cantidad_graficos % 3 != 0 else 0)
                columnas = 3

                fig, axes = plt.subplots(filas, columnas, figsize=(15, 5 * filas))
                axes = axes.flatten()  # Asegura que podemos iterar sin importar la forma

                for i, posicion in enumerate(range(inicio, fin + 1)):
                    numeros_posicion = [
                        row[posicion - 1] for row in ws.iter_rows(min_row=2, max_col=8, values_only=True)
                        if row[posicion - 1] is not None
                    ]

                    if numeros_posicion:
                        conteo = Counter(numeros_posicion)
                        numeros = list(conteo.keys())
                        frecuencias = list(conteo.values())

                        ax = axes[i]
                        ax.bar(numeros, frecuencias, color='skyblue')
                        ax.set_title(f"Posición {posicion}", fontsize=14)
                        ax.set_xlabel("Números", fontsize=10)
                        ax.set_ylabel("Frecuencia", fontsize=10)
                        barras = ax.bar(numeros, frecuencias, color='skyblue')
                        ax.bar_label(barras, fontsize=10, padding=3,color='red')
                        ax.grid(axis='y', linestyle='--', alpha=0.7)
                        ax.set_xticks(numeros)
                        ax.set_xticklabels(numeros, rotation=45)

                # Quitar subplots vacíos si hay
                for j in range(i + 1, len(axes)):
                    fig.delaxes(axes[j])

                plt.tight_layout()
                plt.show()
            else:
                self._mostrar_analisis("No hay datos en Excel para generar estadísticas por posición")
        except FileNotFoundError:
            self._mostrar_analisis("Archivo Excel no encontrado")        

class Bola:
    def __init__(self, canvas, centro_x, centro_y, radio_area):
        self.canvas = canvas
        self.radio_area = radio_area
        self.centro_x = centro_x
        self.centro_y = centro_y
        self.x = centro_x
        self.y = centro_y
        self.dx = random.uniform(-3, 3)
        self.dy = random.uniform(-3, 3)
        self.id = canvas.create_oval(self.x - 10, self.y - 10, self.x + 10, self.y + 10, fill="red")

    def mover(self):
        self.x += self.dx
        self.y += self.dy

        # Rebote si toca el borde de la tómbola (circular)
        dist_x = self.x - self.centro_x
        dist_y = self.y - self.centro_y
        distancia = math.sqrt(dist_x**2 + dist_y**2)
        if distancia + 10 >= self.radio_area:  # +10 por el radio de la bola
            # Rebote: invertir dirección con algo de aleatoriedad
            angulo = math.atan2(dist_y, dist_x)
            normal_x = math.cos(angulo)
            normal_y = math.sin(angulo)
            dot = self.dx * normal_x + self.dy * normal_y
            self.dx -= 2 * dot * normal_x
            self.dy -= 2 * dot * normal_y
            # Pequeño cambio aleatorio
            self.dx += random.uniform(-1, 1)
            self.dy += random.uniform(-1, 1)

        # Mover la bola en el canvas
        self.canvas.coords(self.id, self.x - 10, self.y - 10, self.x + 10, self.y + 10)

def animar_bolas_erraticas(canvas, bolas, delay):
    for bola in bolas:
        bola.mover()
    canvas.after(delay, animar_bolas_erraticas, canvas, bolas, delay)
        

def mostrar_animacion_tombola():
    ventana_animacion = tk.Toplevel(root)
    ventana_animacion.title("Tómbola con Bolas Erráticas")
    ventana_animacion.geometry("400x400")

    canvas = tk.Canvas(ventana_animacion, width=400, height=400, bg="white")
    canvas.pack()

    # Dibujar tómbola
    canvas.create_oval(50, 50, 350, 350, outline="black", width=2)

    # Parámetros
    centro_x, centro_y = 200, 200
    radio_area = 150
    num_bolas = 40
    delay = 50

    # Crear bolas erráticas
    bolas = [Bola(canvas, centro_x, centro_y, radio_area) for _ in range(num_bolas)]

    # Iniciar animación
    animar_bolas_erraticas(canvas, bolas, delay) 
    delay = 5000
    canvas.after(delay, lambda: ventana_animacion.destroy())  # Cerrar ventana después de 5 segundos
    # Crear un nuevo canvas para la animación       
 
    

# Configurar la ventana principal
root = tk.Tk()
root.title("Tombola Multijuego")
root.geometry("1100x600")

# Crear un Canvas para permitir el desplazamiento
canvas = tk.Canvas(root)
canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

# Agregar una barra de desplazamiento vertical
scrollbar = tk.Scrollbar(root, orient=tk.VERTICAL, command=canvas.yview)
scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

# Configurar el Canvas para usar la barra de desplazamiento
canvas.configure(yscrollcommand=scrollbar.set)
canvas.bind('<Configure>', lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

# Crear un marco dentro del Canvas para contener los widgets
frame_contenido = tk.Frame(canvas)
canvas.create_window((0, 0), window=frame_contenido, anchor="nw")

# Crear un marco para organizar las tómbolas horizontalmente
frame_tombolas = tk.Frame(frame_contenido)
frame_tombolas.pack(pady=10)

# Crear las tres tómbolas dentro del marco horizontal
tombola_principal = Tombola(frame_tombolas, rango=(1, 40), columnas=10, max_sorteos=6, 
                            titulo="TOMBOLLA PRINCIPAL (1-40)")
tombola_principal.marco.pack(side=tk.LEFT, padx=10)

tombola_bono1 = Tombola(frame_tombolas, rango=(1, 12), columnas=6, max_sorteos=1, 
                        titulo="BOLA BONO 1 (1-12)")
tombola_bono1.marco.pack(side=tk.LEFT, padx=10)

tombola_bono2 = Tombola(frame_tombolas, rango=(1, 15), columnas=5, max_sorteos=1, 
                        titulo="BOLA BONO 2 (1-15)")
tombola_bono2.marco.pack(side=tk.LEFT, padx=10)

def generar_todos_en_todas():
    """
    Genera todos los números en todas las tómbolas.
    """
    mostrar_animacion_tombola()  # Mostrar la animación
    root.after(1500, lambda: [tombola_principal.generar_todos(),
                              tombola_bono1.generar_todos(),
                              tombola_bono2.generar_todos()])

btn_generar_todos_global = tk.Button(frame_contenido, text="Generar Todos en Todas", command=generar_todos_en_todas,
                                     font=("Arial", 12), bg=COLORS["button"], fg=COLORS["text"])
btn_generar_todos_global.pack(pady=10)

def reiniciar_todas():
    """
    Reinicia todas las tómbolas.
    """
    tombola_principal.reiniciar()
    tombola_bono1.reiniciar()
    tombola_bono2.reiniciar()

btn_reiniciar = tk.Button(frame_contenido, text="Reiniciar Concurso", command=reiniciar_todas,
                          font=("Arial", 12), bg=COLORS["background"], fg=COLORS["text"])
btn_reiniciar.pack(pady=10)

root.mainloop()