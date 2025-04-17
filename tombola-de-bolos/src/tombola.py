import tkinter as tk
import random
import openpyxl
from collections import Counter
from ruta import Ruta
import matplotlib.pyplot as plt
from analisis_estadisticas import (
    predecir_en_base_a_excel,
    predecir_posicion,
    analizar_numeros,
    graficar_frecuencias,
    generar_estadisticas,
    estadisticas_por_posicion,
    estadisticas_por_posicion_r
)

EXCEL_FILE = "numeros_generados.xlsx"
COLORS = {
    "background": "#4CAF50",
    "button": "#FF9800",
    "highlight": "#f44336",
    "text": "white",
    "default": "black"
}

class Tombola:
    """
    Clase que representa una tómbola con números generados aleatoriamente.
    """
    def __init__(self, master, rango, columnas, max_sorteos, titulo):
        """
        Inicializa la tómbola con los parámetros dados.

        Args:
            master (tk.Frame): El marco principal donde se colocará la tómbola.
            rango (tuple): Rango de números (min, max).
            columnas (int): Número de columnas para mostrar los números.
            max_sorteos (int): Número máximo de sorteos permitidos.
            titulo (str): Título de la tómbola.
        """
        self.master = master
        self.min_num, self.max_num = rango
        self.columnas = columnas
        self.max_sorteos = max_sorteos
        self.titulo = titulo

        self.numeros = list(range(self.min_num, self.max_num + 1))
        self.sacados = []
        self.prediccion = None
        self.prediccion_aciertos = 0

        self.crear_widgets()
        self.actualizar_bombo()

    def guardar_en_excel(self):
        """
        Guarda los números generados en un archivo Excel, todos en la misma hoja.
        """
        archivo = Ruta.obtener_ruta_excel(EXCEL_FILE)

        try:
            wb = openpyxl.load_workbook(archivo)
        except FileNotFoundError:
            wb = openpyxl.Workbook()

        if "Resultados" not in wb.sheetnames:
            ws = wb.create_sheet("Resultados")
            ws.append(["Posición 1", "Posición 2", "Posición 3", "Posición 4", "Posición 5", "Posición 6", "BOLA BONO 1", "BOLA BONO 2"])
        else:
            ws = wb["Resultados"]

        fila = ws.max_row + 1
        if self.titulo == "TOMBOLLA PRINCIPAL (1-40)":
            for i, numero in enumerate(self.sacados):
                ws.cell(row=fila, column=i + 1, value=numero)
        elif self.titulo == "BOLA BONO 1 (1-12)" and self.sacados:
            ws.cell(row=fila - 1, column=7, value=self.sacados[0])
        elif self.titulo == "BOLA BONO 2 (1-15)" and self.sacados:
            ws.cell(row=fila - 1, column=8, value=self.sacados[0])

        wb.save(archivo)

    def predecir_en_base_a_excel(self):
        prediccion = predecir_en_base_a_excel()
        if prediccion is not None:
            self.ultimo_numero.set(f"Predicción basada en Excel: {prediccion}")
        else:
            self.ultimo_numero.set("No hay datos en Excel para predecir")

    def predecir_posicion(self, posicion):
        prediccion = predecir_posicion(posicion)
        if prediccion is not None:
            self.ultimo_numero.set(f"Predicción posición {posicion}: {prediccion}")
        else:
            self.ultimo_numero.set(f"No hay datos para predecir posición {posicion}")

    def analizar_numeros(self):
        calientes, frios = analizar_numeros()
        if calientes or frios:
            resultado = f"Calientes: {', '.join([str(num) for num, _ in calientes])}\n"
            resultado += f"Fríos: {', '.join([str(num) for num, _ in frios])}"
            self.texto_analisis.config(state=tk.NORMAL)
            self.texto_analisis.delete(1.0, tk.END)
            self.texto_analisis.insert(tk.END, resultado)
            self.texto_analisis.config(state=tk.DISABLED)
        else:
            self.texto_analisis.config(state=tk.NORMAL)
            self.texto_analisis.delete(1.0, tk.END)
            self.texto_analisis.insert(tk.END, "No hay datos para análisis")
            self.texto_analisis.config(state=tk.DISABLED)

    def _mostrar_analisis(self,resultado):
        """Muestra el análisis de números en el Text widget."""

        self.texto_analisis.config(state=tk.NORMAL)
        self.texto_analisis.delete(1.0, tk.END)
        self.texto_analisis.insert(tk.END, resultado)
        if resultado:
            self.texto_analisis.config(state=tk.DISABLED)
        else:       
            self.texto_analisis.config(state=tk.NORMAL)
            self.texto_analisis.delete(1.0, tk.END)
       
        self.texto_analisis.config(state=tk.DISABLED)    

    def analizar_numeros_principal(self):
        calientes, frios = analizar_numeros(1, 6)
        if calientes or frios:
            resultado = f"Calientes (Principal): {', '.join([str(num) for num, _ in calientes])}\n"
            resultado += f"Fríos (Principal): {', '.join([str(num) for num, _ in frios])}"
            self._mostrar_analisis(resultado)
        else:
            self._mostrar_analisis("No hay datos para análisis en TOMBOLLA PRINCIPAL")

    def analizar_numeros_bono1(self):
        calientes, frios = analizar_numeros(7, 7)
        if calientes or frios:
            resultado = f"Calientes (Bono 1): {', '.join([str(num) for num, _ in calientes])}\n"
            resultado += f"Fríos (Bono 1): {', '.join([str(num) for num, _ in frios])}"
            self._mostrar_analisis(resultado)
        else:
            self._mostrar_analisis("No hay datos para análisis en BOLA BONO 1")

    def analizar_numeros_bono2(self):
        calientes, frios = analizar_numeros(8, 8)
        if calientes or frios:
            resultado = f"Calientes (Bono 2): {', '.join([str(num) for num, _ in calientes])}\n"
            resultado += f"Fríos (Bono 2): {', '.join([str(num) for num, _ in frios])}"
            self._mostrar_analisis(resultado)
        else:
            self._mostrar_analisis("No hay datos para análisis en BOLA BONO 2")

    def generar_todos(self):
        """
        Genera todos los números automáticamente.
        """
        while self.numeros and len(self.sacados) < self.max_sorteos:
            numero = random.choice(self.numeros)
            self.sacados.append(numero)
            self.numeros.remove(numero)

        self.sacados.sort()
        self.actualizar_bombo()
        self._actualizar_texto_sacados()
        self.ultimo_numero.set("Sorteo completo")
        self.btn_sacar.config(text="SORTEO COMPLETO", state=tk.DISABLED)
        self.btn_generar_todos.config(state=tk.DISABLED)
        self.guardar_en_excel()

    def crear_widgets(self):
        """
        Crea los widgets de la interfaz para la tómbola.
        """
        self.marco = tk.Frame(self.master, bd=2, relief=tk.GROOVE)
        self.marco.pack(pady=10, padx=10, fill=tk.BOTH)

        tk.Label(self.marco, text=self.titulo, font=("Arial", 12, "bold")).pack(pady=5)

        self.frame_bombo = tk.Frame(self.marco)
        self.frame_bombo.pack(pady=5)

        self.frame_numeros = tk.Frame(self.frame_bombo)
        self.frame_numeros.pack()

        self.btn_sacar = tk.Button(self.marco, text="Sacar número", command=self.iniciar_sorteo,
                                   font=("Arial", 10), bg=COLORS["background"], fg=COLORS["text"])
        self.btn_sacar.pack(pady=5)

        self.btn_generar_todos = tk.Button(self.marco, text="Generar todos", command=self.generar_todos,
                                           font=("Arial", 10), bg=COLORS["button"], fg=COLORS["text"])
        self.btn_generar_todos.pack(pady=5)

        self.ultimo_numero = tk.StringVar()
        tk.Label(self.marco, textvariable=self.ultimo_numero, font=("Arial", 14), fg=COLORS["highlight"]).pack(pady=5)

        self.lista_sacados = tk.Text(self.marco, width=30, height=4, wrap=tk.WORD, state=tk.DISABLED)
        self.lista_sacados.pack(pady=5)

        # Botón para predecir un número basado en Excel
        self.btn_predecir_excel = tk.Button(self.marco, text="Predecir con Excel", 
                                            command=self.predecir_en_base_a_excel,
                                            font=("Arial", 10), bg="#FFC107", fg="black")
        self.btn_predecir_excel.pack(pady=5)

        if self.titulo == "TOMBOLLA PRINCIPAL (1-40)":
        # Botones para predecir cada posición
            for i in range(1, 7):  # 6 posiciones
                btn_predecir_posicion = tk.Button(self.marco, text=f"Predecir Posición {i}", 
                                                command=lambda i=i: self.predecir_posicion(i),
                                                font=("Arial", 10), bg="#FFC107", fg="black")
                btn_predecir_posicion.pack(pady=2)

        # Botón para analizar números calientes y fríos
            btn_analizar = tk.Button(self.marco, text="Analizar Números", 
                                    command=self.analizar_numeros,
                                    font=("Arial", 10), bg="#4CAF50", fg="white")
            btn_analizar.pack(pady=5)

        # Botón para analizar TOMBOLLA PRINCIPAL
            btn_analizar_principal = tk.Button(self.marco, text="Analizar Principal", 
                                            command=self.analizar_numeros_principal,
                                            font=("Arial", 10), bg="#4CAF50", fg="white")
            btn_analizar_principal.pack(pady=5)
        
        if  self.titulo == "BOLA BONO 1 (1-12)":
        # Botón para analizar BOLA BONO 1
            btn_analizar_bono1 = tk.Button(self.marco, text="Analizar Bono 1", 
                                            command=self.analizar_numeros_bono1,
                                            font=("Arial", 10), bg="#4CAF50", fg="white")
            btn_analizar_bono1.pack(pady=5)
        if self.titulo == "BOLA BONO 2 (1-15)":
        # Botón para analizar BOLA BONO 2
            btn_analizar_bono2 = tk.Button(self.marco, text="Analizar Bono 2", 
                                            command=self.analizar_numeros_bono2,
                                            font=("Arial", 10), bg="#4CAF50", fg="white")
            btn_analizar_bono2.pack(pady=5)

        # Text widget para mostrar el análisis de números
        self.texto_analisis = tk.Text(self.marco, width=40, height=5, wrap=tk.WORD, state=tk.DISABLED)
        self.texto_analisis.pack(pady=5)

        btn_estadisticas = tk.Button(self.marco, text="Generar Estadísticas", 
                                     command=self.generar_estadisticas,
                                     font=("Arial", 10), bg="#4CAF50", fg="white")
        btn_estadisticas.pack(pady=5)

        btn_estadisticas_por_posicion = tk.Button(self.marco, text="Generar Estadísticas por Posición", 
                                                  command=self.generar_estadisticas_por_posicion,
                                                  font=("Arial", 10), bg="#4CAF50", fg="white")
        btn_estadisticas_por_posicion.pack(pady=5)

        btn_estadisticas_por_posicion = tk.Button(self.marco, text=" Estadísticas por Posición r", 
                                                  command=self.estadisticas_por_posicion,
                                                  font=("Arial", 10), bg="#4CAF50", fg="white")
        btn_estadisticas_por_posicion.pack(pady=5)

        

    def actualizar_bombo(self):
        """
        Actualiza los números en el bombo.
        """
        for widget in self.frame_numeros.winfo_children():
            widget.destroy()

        self.numeros.sort()
        for i, num in enumerate(self.numeros):
            fila, columna = divmod(i, self.columnas)
            lbl = tk.Label(self.frame_numeros, text=str(num), width=3, relief=tk.RIDGE,
                           font=("Arial", 10), bg="white")
            lbl.grid(row=fila, column=columna, padx=2, pady=2)

    def iniciar_sorteo(self):
        """
        Inicia el sorteo de un número.
        """
        if self.numeros and len(self.sacados) < self.max_sorteos:
            self.btn_sacar.config(state=tk.DISABLED)
            numero = random.choice(self.numeros)
            self.master.after(1000, lambda: self.completar_sorteo(numero))

    def completar_sorteo(self, numero):
        """
        Completa el sorteo y verifica si la predicción se cumplió.
        """
        if numero in self.numeros:
            self.numeros.remove(numero)
        self.sacados.append(numero)
        self.actualizar_bombo()

        if numero == self.prediccion:
            self.prediccion_aciertos += 1
            self.ultimo_numero.set(f"¡Acierto! Predicción: {self.prediccion} (Aciertos: {self.prediccion_aciertos})")
        else:
            self.ultimo_numero.set(f"Último: {numero} (Aciertos: {self.prediccion_aciertos})")

        self._actualizar_texto_sacados()

        if len(self.sacados) >= self.max_sorteos or not self.numeros:
            self.btn_sacar.config(text="SORTEO COMPLETO", state=tk.DISABLED)

    def _actualizar_texto_sacados(self):
        """
        Actualiza el widget de texto con los números sacados.
        """
        self.lista_sacados.config(state=tk.NORMAL)
        self.lista_sacados.delete(1.0, tk.END)
        columnas = 5
        for i, num in enumerate(self.sacados):
            if i > 0 and i % columnas == 0:
                self.lista_sacados.insert(tk.END, "\n")
            self.lista_sacados.insert(tk.END, f"{num:3} ")
        self.lista_sacados.config(state=tk.DISABLED)

    def reiniciar(self):
        """Reinicia la tómbola a su estado inicial."""
        self.numeros = list(range(self.min_num, self.max_num + 1))
        self.sacados = []
        self.prediccion = None
        self.prediccion_aciertos = 0
        self.actualizar_bombo()
        self.lista_sacados.config(state=tk.NORMAL)
        self.lista_sacados.delete(1.0, tk.END)  # Limpiar el widget Text
        self.lista_sacados.config(state=tk.DISABLED)
        self.ultimo_numero.set("")
        self.btn_sacar.config(text="Sacar número", state=tk.NORMAL)
        self.btn_generar_todos.config(state=tk.NORMAL)

    def predecir_numero(self):
        """Realiza una predicción de un número aleatorio de los números restantes."""
        if self.numeros:
            self.prediccion = random.choice(self.numeros)
            self.ultimo_numero.set(f"Predicción: {self.prediccion}")
        else:
            self.ultimo_numero.set("No hay números para predecir")

    def generar_estadisticas(self):
        if not generar_estadisticas(self.titulo):
            self._mostrar_analisis("No hay posiciones definidas para esta tómbola.")

    def generar_estadisticas_por_posicion(self):
        estadisticas_por_posicion(self.titulo, self._mostrar_analisis)

    def estadisticas_por_posicion(self):
        estadisticas_por_posicion_r(self.titulo, self._mostrar_analisis)

    