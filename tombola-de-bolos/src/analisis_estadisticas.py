import openpyxl
import matplotlib.pyplot as plt
import random
from collections import Counter
from ruta import Ruta

EXCEL_FILE = Ruta.EXCEL_FILE

def predecir_en_base_a_excel():
    archivo = Ruta.obtener_ruta_excel(EXCEL_FILE)
    try:
        wb = openpyxl.load_workbook(archivo)
        if "Resultados" in wb.sheetnames:
            ws = wb["Resultados"]
            numeros_generados = [
                num for row in ws.iter_rows(min_row=2, max_col=8, values_only=True) for num in row if num is not None
            ]
            if numeros_generados:
                conteo = Counter(numeros_generados)
                menos_frecuentes = sorted(conteo.items(), key=lambda x: x[1])[:5]
                return random.choice([num for num, _ in menos_frecuentes])
    except FileNotFoundError:
        pass
    return None

def predecir_posicion(posicion):
    archivo = Ruta.obtener_ruta_excel(EXCEL_FILE)
    try:
        wb = openpyxl.load_workbook(archivo)
        if "Resultados" in wb.sheetnames:
            ws = wb["Resultados"]
            numeros_posicion = [
                row[posicion - 1] for row in ws.iter_rows(min_row=2, max_col=8, values_only=True)
                if row[posicion - 1] is not None
            ]
            if numeros_posicion:
                conteo = Counter(numeros_posicion)
                menos_frecuentes = sorted(conteo.items(), key=lambda x: x[1])[:5]
                return random.choice([num for num, _ in menos_frecuentes])
    except FileNotFoundError:
        pass
    return None

def analizar_numeros(col_inicio=1, col_fin=8):
    archivo = Ruta.obtener_ruta_excel(EXCEL_FILE)
    try:
        wb = openpyxl.load_workbook(archivo)
        if "Resultados" in wb.sheetnames:
            ws = wb["Resultados"]
            numeros_generados = []
            for row in ws.iter_rows(min_row=2, max_col=col_fin, values_only=True):
                numeros_generados.extend(filter(None, row[col_inicio-1:col_fin]))
            if numeros_generados:
                conteo = Counter(numeros_generados)
                calientes = sorted(conteo.items(), key=lambda x: x[1], reverse=True)[:5]
                frios = sorted(conteo.items(), key=lambda x: x[1])[:5]
                return calientes, frios
    except FileNotFoundError:
        pass
    return [], []

def graficar_frecuencias(titulo, posiciones):
    archivo = Ruta.obtener_ruta_excel(EXCEL_FILE)
    try:
        wb = openpyxl.load_workbook(archivo)
        if "Resultados" in wb.sheetnames:
            ws = wb["Resultados"]
            fig, axes = plt.subplots(len(posiciones), 1, figsize=(15, 5 * len(posiciones)))
            if len(posiciones) == 1:
                axes = [axes]
            for ax, posicion in zip(axes, posiciones):
                numeros = [
                    row[posicion - 1] for row in ws.iter_rows(min_row=2, max_col=8, values_only=True)
                    if row[posicion - 1] is not None
                ]
                if numeros:
                    conteo = Counter(numeros)
                    nums = list(conteo.keys())
                    freqs = list(conteo.values())
                    barras = ax.bar(nums, freqs, color='skyblue')
                    ax.bar_label(barras, fontsize=10, padding=3, color='red')
                    ax.set_title(f"Frecuencia de Números - Posición {posicion}", fontsize=14)
                    ax.set_xlabel("Números", fontsize=12)
                    ax.set_ylabel("Frecuencia", fontsize=12)
                    ax.set_xticks(nums)
                    ax.set_xticklabels(nums, rotation=45)
                    ax.grid(axis='y', linestyle='--', alpha=0.7)
            plt.tight_layout()
            plt.show()
    except FileNotFoundError:
        pass

def generar_estadisticas(titulo):
    posiciones_tombola = {
        "TOMBOLLA PRINCIPAL (1-40)": range(1, 7),
        "BOLA BONO 1 (1-12)": range(7, 8),
        "BOLA BONO 2 (1-15)": range(8, 9)
    }
    posiciones = posiciones_tombola.get(titulo, [])
    if posiciones:
        graficar_frecuencias(titulo, posiciones)
        return True
    return False

def estadisticas_por_posicion(titulo, mostrar_analisis_callback):
    archivo = Ruta.obtener_ruta_excel(EXCEL_FILE)
    posicion_tombola = {
    "GENERAL": (1, 7),
    "BOLA BONO 1": (7, 8),
    "BOLA BONO 2": (8, 9)
    }
    try:
        wb = openpyxl.load_workbook(archivo)
        if "Resultados" in wb.sheetnames:
            ws = wb["Resultados"]
            if titulo == "TOMBOLLA PRINCIPAL (1-40)":
                position_rango = posicion_tombola["GENERAL"]
            elif titulo == "BOLA BONO 1 (1-12)":
                position_rango = posicion_tombola["BOLA BONO 1"]
            elif titulo == "BOLA BONO 2 (1-15)":
                position_rango = posicion_tombola["BOLA BONO 2"]
            else:   
                position_rango = (1, 9)
            # Iterar por cada posición (columna)
            inicio, fin = position_rango
            for posicion in range( inicio, fin):  # Columnas 1 a 8
                numeros_posicion = [
                    row[posicion - 1] for row in ws.iter_rows(min_row=2, max_col=8, values_only=True)
                    if row[posicion - 1] is not None
                ]
                
                if numeros_posicion:
                    # Contar la frecuencia de cada número
                    from collections import Counter
                    conteo = Counter(numeros_posicion)
                    
                    # Preparar datos para el gráfico
                    numeros = list(conteo.keys())
                    frecuencias = list(conteo.values())
                    
                    # Crear el gráfico de barras
                    plt.figure(figsize=(10, 6))
                    plt.bar(numeros, frecuencias, color='skyblue')
                    plt.title(f"Frecuencia de Números - Posición {posicion}", fontsize=16)
                    plt.xlabel("Números", fontsize=12)
                    plt.ylabel("Frecuencia", fontsize=12)
                    plt.xticks(numeros,rotation=45)
                    plt.tight_layout()
                    
                    # Mostrar el gráfico
                    plt.show()
        else:
            mostrar_analisis_callback("No hay datos en Excel para generar estadísticas por posición")
    except FileNotFoundError:
        mostrar_analisis_callback("Archivo Excel no encontrado")
def estadisticas_por_posicion_r(titulo,mostrar_analisis_callback):
        """Genera estadísticas por posición de cada número y muestra gráficos."""
        archivo = Ruta.obtener_ruta_excel(EXCEL_FILE)
        posicion_tombola = {
        "GENERAL": (1, 6),
        "BOLA BONO 1": (7, 7),
        "BOLA BONO 2": (8, 8)
        }
        try:
            wb = openpyxl.load_workbook(archivo)
            if "Resultados" in wb.sheetnames:
                ws = wb["Resultados"]
                if titulo == "TOMBOLLA PRINCIPAL (1-40)":
                    position_rango = posicion_tombola["GENERAL"]
                elif titulo == "BOLA BONO 1 (1-12)":
                    position_rango = posicion_tombola["BOLA BONO 1"]
                elif titulo == "BOLA BONO 2 (1-15)":
                    position_rango = posicion_tombola["BOLA BONO 2"]
                else:   
                    position_rango = (1, 9)
                # Iterar por cada posición (columna)
                inicio, fin = position_rango
               # Calcular cuántos gráficos vamos a hacer
                cantidad_graficos = fin - inicio + 1

                # Calcular filas y columnas para distribuir bien los subplots
                filas = (cantidad_graficos // 3) + (1 if cantidad_graficos % 3 != 0 else 0)
                columnas = 3

                fig, axes = plt.subplots(filas, columnas, figsize=(15, 5 * filas))
                axes = axes.flatten()  # Asegura que podemos iterar sin importar la forma

                for i, posicion in enumerate(range(inicio, fin + 1)):
                    numeros_posicion = [
                        row[posicion - 1] for row in ws.iter_rows(min_row=2, max_col=8, values_only=True)
                        if row[posicion - 1] is not None
                    ]

                    if numeros_posicion:
                        conteo = Counter(numeros_posicion)
                        numeros = list(conteo.keys())
                        frecuencias = list(conteo.values())

                        ax = axes[i]
                        ax.bar(numeros, frecuencias, color='skyblue')
                        ax.set_title(f"Posición {posicion}", fontsize=14)
                        ax.set_xlabel("Números", fontsize=10)
                        ax.set_ylabel("Frecuencia", fontsize=10)
                        barras = ax.bar(numeros, frecuencias, color='skyblue')
                        ax.bar_label(barras, fontsize=10, padding=3,color='red')
                        ax.grid(axis='y', linestyle='--', alpha=0.7)
                        ax.set_xticks(numeros)
                        ax.set_xticklabels(numeros, rotation=45)

                # Quitar subplots vacíos si hay
                for j in range(i + 1, len(axes)):
                    fig.delaxes(axes[j])

                plt.tight_layout()
                plt.show()
            else:
                mostrar_analisis_callback("No hay datos en Excel para generar estadísticas por posición")
        except FileNotFoundError:
            mostrar_analisis_callback("Archivo Excel no encontrado")
